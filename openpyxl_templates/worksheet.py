from collections import deque
from itertools import chain

from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from openpyxl_templates.exceptions import OpenpyxlTemplateCellException, CellExceptions
from openpyxl_templates.style import CellStyle
from openpyxl_templates.workbook import StyleSet

MAX_COLUMN_INDEX = column_index_from_string("XFD")


class ExcelRow:
    column_attrs = None
    row_number = None

    def __init__(self, column_attrs, row_number=None):
        self.column_attrs = column_attrs
        self.row_number = row_number

        for attr in self.column_attrs:
            setattr(self, attr, None)

    def __str__(self):
        return "Row %d: {%s}" % (
            self.row_number,
            ", ".join(("%s: '%s'" % (attr, getattr(self, attr)) for attr in self.column_attrs))
        )


class RowExceptionPolicy:
    IGNORE_ROW = "IGNORE_ROW"
    RAISE_EXCEPTION = "RAISE_EXCEPTION"
    RETURN_EXCEPTION = "RETURN_EXCEPTION"


class SheetTemplate:
    name = None
    title = None
    description = None

    columns = None
    # style = None
    # header_style = None
    # title_style = None
    # description_style = None

    styles = None
    title_style = StyleSet.DEFAULT_TITLE_STYLE
    heading_style = StyleSet.DEFAULT_HEADER_STYLE


    freeze_header = True
    hide_excess_columns = True
    format_as_table = True
    table_style = None
    empty_row_count = 1

    row_exception_policy = RowExceptionPolicy.RAISE_EXCEPTION

    def __init__(self, name=None, title=None, description=None, columns=None, styles=None, freeze_header=None,
                 hide_excess_columns=None, format_as_table=None, table_style=None, empty_row_count=None,
                 row_exception_policy=None):

        self.name = name or self.name
        self.title = title or self.title
        self.description = description or self.description
        self.columns = columns or self.columns or []
        self.styles = list(styles or self.styles or [])

        for column in self.columns:
            for style in column.styles:
                self.styles.append(style)

        self.freeze_header = freeze_header if freeze_header is not None else self.freeze_header
        self.hide_excess_columns = hide_excess_columns if hide_excess_columns is not None else self.hide_excess_columns
        self.format_as_table = format_as_table if format_as_table is not None else self.format_as_table
        self.table_style = table_style or table_style
        self.empty_row_count = empty_row_count if empty_row_count is not None else self.empty_row_count

        self.row_exception_policy = row_exception_policy or self.row_exception_policy

        self._column_attrs = list((column.object_attr for column in self.columns))

    def read_rows(self, worksheet):
        row_number = 0
        for raw_row in worksheet:
            row_number += 1
            row = self.create_empty_row(row_number=row_number)

            que = deque(raw_row)
            errors = []
            for column in self.columns:
                cell = que.popleft() if que else None

                try:
                    value = column.from_excel_with_blank_check(cell)
                except OpenpyxlTemplateCellException as cell_exception:
                    value = cell_exception
                    errors.append(cell_exception)

                column.set_value_to_object(row, value)

            if errors:
                if self.row_exception_policy == RowExceptionPolicy.IGNORE_ROW:
                    continue

                row_exception = CellExceptions(cell_exceptions=errors)
                if self.row_exception_policy == RowExceptionPolicy.RETURN_EXCEPTION:
                    yield row_exception
                else:
                    raise row_exception

            yield row

    def write_headers(self, worksheet, style_set):
        headers = []
        for column in self.columns:
            header = WriteOnlyCell(ws=worksheet, value=column.header)
            style = style_set[column.header_style]
            if style:
                header.style = style
            headers.append(header)

        worksheet.append(headers)
        first_cell = headers[0]
        return first_cell

    def write_rows(self, worksheet, style_set, objects):
        styles = tuple(style_set[column.row_style] for column in self.columns)
        data_validations = tuple(column.data_validation for column in self.columns)
        rows = (
            (
                column.create_cell(worksheet, obj=obj)
                for column in self.columns
            ) for obj in objects
        )
        empty_rows = (
            (
                column.create_cell(worksheet, obj=None)
                for column in self.columns
            ) for i in range(0, self.empty_row_count)
        )

        row_count = 0
        first_cell = None
        cells = tuple()
        for row in chain(rows, empty_rows):
            row_count += 1

            cells = tuple(row)
            worksheet.append(cells)

            if not first_cell:
                first_cell = cells[0]
                if self.freeze_header:
                    worksheet.freeze_panes = first_cell

            """
                Styling is separated from creation since data_validation
                must be applied after appending to worksheet
            """
            for cell, style, data_validation in zip(cells, styles, data_validations):
                if style:
                    cell.style = style
                if data_validation:
                    data_validation.add(cell)

        last_cell = cells[-1]

        return first_cell, last_cell

    def style_columns(self, worksheet, style_set):
        for index, column in enumerate(self.columns):
            column_dimension = worksheet.column_dimensions[get_column_letter(index + 1)]
            if column.width is not None:
                column_dimension.width = column.width

            row_style = style_set[column.row_style]
            if row_style:
                column_dimension.style = row_style

            column_dimension.hidden = column.hidden

        if self.hide_excess_columns:
            for i in range(len(self.columns) + 1, MAX_COLUMN_INDEX + 1):
                worksheet.column_dimensions[get_column_letter(i)].hidden = True

    def create_empty_row(self, row_number):
        return ExcelRow(self._column_attrs, row_number=row_number)

    def write_title(self, worksheet):
        # if self.title:
        #     worksheet.append((self.title_style.style_cell(WriteOnlyCell(worksheet, value=self.title)),))
        #     if not self.description:
        #         worksheet.append((None,))
        pass

    def write_description(self, worksheet):
        # if self.description:
        #     worksheet.append((self.description_style.style_cell(WriteOnlyCell(worksheet, value=self.description)),))
        #     worksheet.append((None,))
        pass

    def write(self, worksheet, style_set, objects):
        # worksheet = self.workbook.create_sheet(self.name)
        self.write_title(worksheet)
        self.write_description(worksheet)
        first_header = self.write_headers(worksheet, style_set)
        first_row_cell, last_row_cell = self.write_rows(worksheet, style_set, objects)
        self.style_columns(worksheet, style_set)

        if self.format_as_table:
            table = Table(
                ref="%s:%s" % (first_header.coordinate, last_row_cell.coordinate),
                displayName=self.name,
                tableStyleInfo=self.table_style
            )
            worksheet.add_table(table)

            # def rebase_column_styles(self):
            #     for column in self.columns:
            #         column.style = CellStyle.merge(self.style, column.style)
            #         column.header_style = CellStyle.merge(self.header_style, column.header_style)
